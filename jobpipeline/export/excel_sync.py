from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from jobpipeline.core.models import CanonicalJob

COLUMNS = [
    "Job ID",
    "Date Collected",
    "Last Seen",
    "Company",
    "Title",
    "Location",
    "Remote",
    "Link",
    "Source(s)",
    "Posted Date",
    "Fit Score",
    "Fit Grade",
    "Fit Notes",
    "Missing Must-have",
    "Status",
    "Notes",
]


class ExcelSync:
    def __init__(self, path: str) -> None:
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def sync(self, jobs: list[CanonicalJob]) -> int:
        if self.path.exists():
            wb = load_workbook(self.path)
            ws = wb["Jobs"] if "Jobs" in wb.sheetnames else wb.create_sheet("Jobs")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Jobs"

        if ws.max_row < 1 or ws.cell(1, 1).value != "Job ID":
            ws.delete_rows(1, ws.max_row)
            ws.append(COLUMNS)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:P1"
            for idx, col in enumerate(COLUMNS, start=1):
                ws.cell(1, idx).font = Font(bold=True)

        job_to_row = {}
        for row_idx in range(2, ws.max_row + 1):
            job_id = ws.cell(row_idx, 1).value
            if job_id:
                job_to_row[str(job_id)] = row_idx

        exported = 0
        for job in jobs:
            row_values = [
                job.job_id,
                job.collected_at,
                job.last_seen,
                job.company,
                job.title,
                job.location_text,
                job.remote_flag,
                job.canonical_url,
                job.source_name,
                job.posted_date or "",
                job.fit_score,
                job.fit_grade,
                job.fit_notes,
                ", ".join(job.missing_must_have),
                job.user_status,
                job.user_notes,
            ]

            if job.job_id in job_to_row:
                ridx = job_to_row[job.job_id]
                current_status = ws.cell(ridx, 15).value or "New"
                current_notes = ws.cell(ridx, 16).value or ""
                for cidx, value in enumerate(row_values, start=1):
                    if cidx in (15, 16):
                        continue
                    ws.cell(ridx, cidx, value)
                ws.cell(ridx, 15, current_status)
                ws.cell(ridx, 16, current_notes)
            else:
                ws.append(row_values)
                ridx = ws.max_row
                ws.cell(ridx, 15, "New")
                ws.cell(ridx, 16, "")

            link_cell = ws.cell(ridx, 8)
            link_cell.hyperlink = str(link_cell.value)
            link_cell.style = "Hyperlink"
            exported += 1

        widths = {
            1: 18,
            2: 20,
            3: 20,
            4: 24,
            5: 36,
            6: 22,
            7: 9,
            8: 60,
            9: 20,
            10: 16,
            11: 10,
            12: 10,
            13: 45,
            14: 26,
            15: 12,
            16: 30,
        }
        for idx, width in widths.items():
            ws.column_dimensions[chr(64 + idx)].width = width

        wb.save(self.path)
        return exported
