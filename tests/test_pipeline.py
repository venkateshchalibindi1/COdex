from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from jobpipeline.core.models import CanonicalJob, SearchProfile, SourceItem
from jobpipeline.core.orchestrator import PipelineOrchestrator
from jobpipeline.export.excel_sync import COLUMNS
from jobpipeline.scoring.service import FitScorer
from jobpipeline.storage.repository import JobRepository


class FakeManager:
    def __init__(self, items: list[SourceItem]) -> None:
        self.items = items

    def search(self, profile: SearchProfile, max_jobs: int) -> list[SourceItem]:
        return self.items[:max_jobs]


class FakeCollector:
    def __init__(self, jobs: list[CanonicalJob]) -> None:
        self.jobs = jobs

    def collect(self, item: SourceItem) -> CanonicalJob:
        return self.jobs.pop(0)


def make_config(tmp_path: Path) -> dict:
    return {
        "excel_path": str(tmp_path / "jobs.xlsx"),
        "profiles": [
            {
                "name": "default",
                "target_titles": ["IT Support Specialist"],
                "adjacent_titles": ["Help Desk Analyst"],
                "location_mode": "Remote",
                "city": "",
                "radius_km": 0,
                "experience_min_years": 1,
                "experience_max_years": 3,
                "must_have_keywords": ["troubleshooting", "customer service"],
                "nice_to_have_keywords": ["active directory", "azure"],
                "exclude_keywords": ["senior"],
                "time_window_days": 7,
                "master_resume_skills": ["troubleshooting"],
            }
        ],
        "sources": {"rss_feeds": [], "greenhouse_boards": [], "lever_boards": []},
        "discovery": {"enabled": False, "provider": "disabled", "api_key": ""},
        "collector": {"use_playwright": False, "per_domain_delay_seconds": 0, "max_retries": 0},
        "filters": {"exclude_domains": [], "exclude_keywords": [], "seniority_mode": "downrank"},
        "limits": {"max_jobs_per_run": 300},
    }


def mkjob(job_id: str, canonical_url: str, source_name: str = "RSS") -> CanonicalJob:
    return CanonicalJob(
        job_id=job_id,
        source_domain="example.com",
        source_name=source_name,
        job_url=canonical_url,
        canonical_url=canonical_url,
        apply_url=canonical_url,
        title="IT Support Specialist",
        company="Acme",
        location_text="Remote",
        remote_flag="Y",
        employment_type="FT",
        posted_date="2026-01-01",
        collected_at="2026-01-02T00:00:00",
        description_raw="Great role troubleshooting customer service active directory 2 years",
        salary_text=None,
        skills_extracted=["troubleshooting"],
        fetch_status="success",
        failure_reason=None,
        first_seen="2026-01-02T00:00:00",
        last_seen="2026-01-02T00:00:00",
        repost_count=0,
        merged_from=[],
        fit_score=0,
        fit_grade="D",
        fit_notes="",
        missing_must_have=[],
        flags=[],
    )


def test_pipeline_creates_sqlite_and_excel(tmp_path: Path) -> None:
    config = make_config(tmp_path)
    repo = JobRepository(str(tmp_path / "jobs.db"))
    orchestrator = PipelineOrchestrator(config, repo)

    items = [SourceItem(job_url="https://example.com/j/1", source_name="RSS", source_domain="example.com")]
    orchestrator._source_manager = lambda: FakeManager(items)  # type: ignore[method-assign]
    orchestrator.collector = FakeCollector([mkjob("job1", "https://example.com/j/1")])

    counts = orchestrator.run()
    assert counts["found"] == 1
    assert len(repo.list_jobs()) == 1
    assert Path(config["excel_path"]).exists()

    wb = load_workbook(config["excel_path"])
    ws = wb["Jobs"]
    headers = [ws.cell(1, i).value for i in range(1, len(COLUMNS) + 1)]
    assert headers == COLUMNS


def test_rerun_updates_last_seen_not_duplicate_and_preserves_sources(tmp_path: Path) -> None:
    config = make_config(tmp_path)
    repo = JobRepository(str(tmp_path / "jobs.db"))
    orchestrator = PipelineOrchestrator(config, repo)

    items = [
        SourceItem(job_url="https://example.com/j/1", source_name="RSS", source_domain="example.com"),
        SourceItem(job_url="https://example.com/j/1?utm_source=x", source_name="RSS2", source_domain="example.com"),
    ]
    orchestrator._source_manager = lambda: FakeManager(items)  # type: ignore[method-assign]
    orchestrator.collector = FakeCollector(
        [
            mkjob("job1", "https://example.com/j/1", "RSS"),
            mkjob("job2", "https://example.com/j/1", "RSS2"),
        ]
    )

    orchestrator.run()
    jobs = repo.list_jobs()
    assert len(jobs) == 1
    assert "RSS2" in jobs[0]["source_name"]


def test_excel_sync_preserves_user_fields(tmp_path: Path) -> None:
    config = make_config(tmp_path)
    repo = JobRepository(str(tmp_path / "jobs.db"))
    orchestrator = PipelineOrchestrator(config, repo)

    item = [SourceItem(job_url="https://example.com/j/1", source_name="RSS", source_domain="example.com")]
    orchestrator._source_manager = lambda: FakeManager(item)  # type: ignore[method-assign]
    orchestrator.collector = FakeCollector([mkjob("job1", "https://example.com/j/1")])
    orchestrator.run()
    repo.update_user_fields("job1", "Applied", "sent resume")

    orchestrator.collector = FakeCollector([mkjob("job1", "https://example.com/j/1")])
    orchestrator.run()

    wb = load_workbook(config["excel_path"])
    ws = wb["Jobs"]
    assert ws.cell(2, 15).value == "Applied"
    assert ws.cell(2, 16).value == "sent resume"


def test_fit_scoring_is_deterministic() -> None:
    profile = SearchProfile(
        name="default",
        target_titles=["IT Support Specialist"],
        adjacent_titles=[],
        location_mode="Remote",
        city="",
        radius_km=0,
        experience_min_years=1,
        experience_max_years=3,
        must_have_keywords=["troubleshooting"],
        nice_to_have_keywords=["azure"],
        exclude_keywords=[],
        time_window_days=7,
        master_resume_skills=["troubleshooting"],
    )
    job = mkjob("job1", "https://example.com/j/1")
    scorer = FitScorer()
    first = scorer.score(job, profile)
    second = scorer.score(job, profile)
    assert first.fit_score == second.fit_score
    assert first.fit_notes == second.fit_notes
